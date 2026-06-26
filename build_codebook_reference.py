#!/usr/bin/env python3
"""
build_codebook_reference.py
===========================
FAITHFUL port of the team reference pipeline
  codebook-reference/count_nps.py            (recount + unigram selection)
  codebook-reference/apply_downscale_algorithm.py  (capped downscale)
adapted to the legal corpus.

Differences from the reference are only the inputs the user asked to keep:
  * candidates = multi-method, lemma-normalised, noun-head-filtered (build_codebook_v4)
  * corpus     = the legally-preprocessed 100 docs, joined as a lemma stream
The METHOD is replicated exactly:
  1. initial Frequency from extraction.
  2. unigram selection: log-frequency KneeLocator cut + keep only 0/1==1
     (informative). 0/1 here = the heuristic 'informative' suggestion, since
     the real run happens after the manual labelling pass.
  3. recount candidate phrases by Aho-Corasick exact (space-padded) matching
     over the concatenated corpus -> new_frequency.
  4. downscale: head-unigram assignment + capped cumulative discount
     cap = max(count/2, count - max_phrase + 1), boundary phrase included.
  5. rank by new_algorithm_frequency, top 3000, annotate, write.

Output: np_codebook_reference.{csv,xlsx}
"""
from __future__ import annotations
import math, csv, collections, sys
import numpy as np
import pandas as pd
import ahocorasick
from kneed import KneeLocator

import build_codebook_v2 as v2
import build_codebook_v4 as v4
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

TOP_N = 3000


def main() -> int:
    terms, total, docf, methods, example, doc_lemmas = v4.extract_candidates()

    # ---- build the codebook frame: phrase, Frequency, 0/1 (heuristic informative) ----
    df = pd.DataFrame({"phrase": terms})
    df["Frequency"] = df["phrase"].map(lambda p: total[p])
    df["0/1"] = df["phrase"].map(lambda p: 1 if v2.classify(p)[1] == "informative" else 0)
    df = df.sort_values("Frequency", ascending=False).reset_index(drop=True)

    # ===== count_nps.py: knee-point unigram selection =====
    unigram_df = df[df["phrase"].apply(lambda x: len(str(x).split()) == 1)].reset_index(drop=True)
    unigram_df["log_frequency"] = unigram_df["Frequency"].apply(lambda x: math.log(x))
    kn = KneeLocator(unigram_df["log_frequency"], unigram_df.index,
                     curve="convex", direction="decreasing", S=2)
    knee = kn.knee
    cut_point = len(unigram_df[unigram_df["log_frequency"] > knee]) if knee is not None else len(unigram_df)
    unigram_df = unigram_df.iloc[:cut_point]
    unigram_df = unigram_df[unigram_df["0/1"] == 1].reset_index(drop=True)
    selected_unigrams = set(unigram_df["phrase"])
    print(f"knee(log_freq)={knee} | unigrams above knee={cut_point} | "
          f"informative head unigrams kept={len(selected_unigrams)}")

    # phrases containing a selected informative unigram
    filtered_df = df[df["phrase"].apply(
        lambda x: any(u in str(x).split() for u in selected_unigrams))].reset_index(drop=True)
    # informative multi-word phrases not already covered
    ni_bigrams_df = df[df.apply(lambda r: len(str(r["phrase"]).split()) != 1 and r["0/1"] == 1,
                                axis=1)].reset_index(drop=True)
    not_computed = set(ni_bigrams_df["phrase"]).difference(set(filtered_df["phrase"]))
    ni_bigrams_df = ni_bigrams_df[ni_bigrams_df["phrase"].isin(not_computed)].reset_index(drop=True)
    filtered_top_df = filtered_df.copy()
    print(f"filtered phrases={len(filtered_top_df)} | extra informative phrases={len(ni_bigrams_df)}")

    # ===== count_nps.py: Aho-Corasick exact recount over the corpus =====
    concatted_doc = " " + " . ".join(" ".join(lem) for lem in doc_lemmas) + " "
    print(f"corpus length (lemma stream chars): {len(concatted_doc)}")
    automaton = ahocorasick.Automaton()
    all_phrases = pd.concat([filtered_top_df["phrase"], ni_bigrams_df["phrase"]]).unique()
    for idx, key in enumerate(all_phrases):
        automaton.add_word(" " + key + " ", (idx, " " + key + " "))
    automaton.make_automaton()
    counts = collections.Counter(orig for _, (_, orig) in automaton.iter(concatted_doc))
    nf = lambda p: counts.get(" " + p + " ", 0)
    filtered_top_df["new_frequency"] = filtered_top_df["phrase"].map(nf)
    ni_bigrams_df["new_frequency"] = ni_bigrams_df["phrase"].map(nf)
    result_df = pd.concat([filtered_top_df, ni_bigrams_df], ignore_index=True)
    result_df = result_df.drop_duplicates("phrase").reset_index(drop=True)

    # ===== apply_downscale_algorithm.py: capped downscale on new_frequency =====
    ndf = result_df[result_df["new_frequency"] > 0].copy()
    ndf["ngram"] = ndf["phrase"].apply(lambda x: tuple(x.split()))
    ngram_count = dict(zip(ndf["ngram"], ndf["new_frequency"]))
    ndf["head_unigram"] = ndf["ngram"].apply(lambda x: max(x, key=lambda y: ngram_count.get((y,), 0)))
    ndf["ngram_length"] = ndf["ngram"].apply(len)
    ndf["new_algorithm_frequency"] = ndf["new_frequency"]
    nf_map = dict(zip(ndf["phrase"], ndf["new_frequency"]))

    for u in ndf[ndf["ngram_length"] == 1]["phrase"]:
        grp = ndf[ndf["head_unigram"] == u].sort_values("new_frequency", ascending=False)
        bigrams = grp[grp["ngram_length"] != 1]
        if bigrams.empty:
            continue
        ucount = nf_map[u]
        max_bg = bigrams["new_frequency"].iloc[0]
        threshold = max(ucount / 2, ucount - max_bg + 1)
        cum = 0
        for f in bigrams["new_frequency"]:
            cum += f
            if cum > threshold:           # include boundary phrase, then stop
                break
        ndf.loc[ndf["phrase"] == u, "new_algorithm_frequency"] = max(ucount - cum, 0)

    ndf = ndf.sort_values("new_algorithm_frequency", ascending=False).reset_index(drop=True)

    # ---- top 3000 + annotate ----
    rows = []
    for rank, r in ndf.head(TOP_N).iterrows():
        term = r["phrase"]
        nonamb, informative, typ, topic, role = v2.classify(term)
        rows.append({
            "rank": rank + 1, "term": term, "doc_freq": docf[term],
            "frequency": int(r["new_frequency"]),
            "adj_frequency": int(r["new_algorithm_frequency"]),
            "methods": ",".join(sorted(methods[term])),
            "nonambiguous": nonamb, "informative": informative,
            "type": typ, "topic": topic, "role": role,
            "notes": "", "example": example.get(term, ""),
        })
    cols = ["rank", "term", "doc_freq", "frequency", "adj_frequency", "methods",
            "nonambiguous", "informative", "type", "topic", "role", "notes", "example"]
    with open("np_codebook_reference.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(rows)
    write_xlsx(rows, cols)
    n_inf = sum(1 for r in rows if r["informative"] == "informative")
    print(f"Wrote np_codebook_reference.{{csv,xlsx}} | {len(rows)} terms | informative {n_inf}")
    return 0


def write_xlsx(rows, cols):
    clean = lambda v: ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v
    wb = Workbook(); ws = wb.active; ws.title = "np_codebook"
    hf = PatternFill("solid", fgColor="1F4E78")
    ws.append([c.upper() for c in cols])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
    for r in rows:
        ws.append([clean(r[c]) for c in cols])
    n = len(rows)

    def dv(col, formula):
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        L = get_column_letter(cols.index(col) + 1)
        d.add(f"{L}2:{L}{n+1}")
    dv("nonambiguous", '"1,0"')
    dv("informative", '"informative,non-informative"')
    dv("type", '"' + ",".join(t for t, _ in v2.TYPES) + '"')
    dv("topic", '"' + ",".join(t for t, _ in v2.TOPICS) + '"')
    dv("role", '"' + ",".join(t for t, _ in v2.ROLES) + '"')
    widths = {"rank": 6, "term": 30, "doc_freq": 9, "frequency": 10, "adj_frequency": 12,
              "methods": 18, "nonambiguous": 13, "informative": 16, "type": 18,
              "topic": 20, "role": 12, "notes": 18, "example": 55}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "C2"
    wb.save("np_codebook_reference.xlsx")


if __name__ == "__main__":
    sys.exit(main())
