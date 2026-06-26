#!/usr/bin/env python3
"""
build_codebook_v3.py
====================
Workflow-compliant rebuild of the NP codebook with the Step-5 frequency
adjustment and Step-6 adjusted-frequency ranking.

Steps implemented here (numbers per the Codebook Construction Workflow):
  2/3  Extract noun chunks (spaCy) over the 100 docs, normalise, dedupe,
       drop stopword-only / numeric chunks.
  4    Compute corpus frequency (token total_freq) and document frequency
       for EVERY candidate term (not just the top 3000).
  5    Frequency adjustment of common unigrams:
         - for each multi-word phrase P, find the highest-frequency unigram
           u* it contains and discount u* by freq(P)
           ("occurrences already represented by a more informative phrase").
         - adj_freq(unigram) = raw_freq - sum(discounts), floored at 0.
         - multi-word phrases keep their own frequency.
  6    Rank all terms by adjusted frequency (tie: document frequency).
  7    Select the top 3000.
  8/9  Pre-annotate (informative / nonambiguous / type / topic / role) via the
       v2 heuristic, for the manual review pass.

Outputs codebook-varshini.{csv,xlsx} (overwrites the v2 raw-ranked version).
"""
from __future__ import annotations
import csv, os, re, collections, sys
import fitz
import spacy

import build_np_codebook as v1          # pdf_text, normalise_chunk
import build_codebook_v2 as v2          # classify, TYPES, TOPICS, ROLES
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

SAMPLE, PDFDIR, TOP_N, MODEL = "sample_100_docs.csv", "documents_pdf", 3000, "en_core_web_md"


# --------------------------------------------------------------------------- #
# Legal-judgment preprocessing
# --------------------------------------------------------------------------- #
# NOTE: Court judgments need DIFFERENT preprocessing than social-media messages.
# Social-media cleaning targets @mentions, #hashtags, emojis, URLs, slang,
# elongated words, retweets. None of that occurs in judgments. Instead, judgment
# text carries: PDF letter-spacing and line-break hyphenation artifacts, an
# Indian Kanoon header/footer, an "equivalent citations" metadata block, reporter
# citation strings (AIR / SCC / SCR ...), bench/author/party metadata labels, and
# page numbers. Those are removed BEFORE noun-chunk extraction so they never
# enter the candidate vocabulary, while substantive legal references
# (Section / Article / Act) are deliberately preserved.

_MONTHS = (r"january|february|march|april|may|june|july|august|september|"
           r"october|november|december")
REPORTERS = (r"AIR|SCC|SCR|SCALE|SCJ|MLJ|KLT|ILR|JT|ALD|ALL|BLJR|Bom|Cal|Mad|Pat|"
             r"Guj|Ker|Raj|Cri\.?\s?L\.?J|Cri\.?\s?L\.?R|Lab\.?\s?I\.?C|ITR")


def preprocess_legal(raw: str) -> str:
    t = raw
    # 1. join line-break / OCR hyphenation splits: 'deci- sion', 'ques-\ntion'
    t = re.sub(r"(\w)-\s+(\w)", r"\1\2", t)
    # 2. collapse PDF letter-spacing: 'e q u i v a l e n t' -> 'equivalent'
    t = re.sub(r"(?:\b[A-Za-z]\b ){3,}\b[A-Za-z]\b",
               lambda m: m.group(0).replace(" ", ""), t)
    t = t.replace("\xad", "")
    # 3. Indian Kanoon header/footer and any URLs
    t = re.sub(r"Indian Kanoon\s*-\s*http\S*", " ", t, flags=re.I)
    t = re.sub(r"https?://\S+", " ", t)
    # 4. 'equivalent citations: ...' metadata block (stop at line end / next field)
    t = re.sub(r"equivalent citations?\s*:.*?(?=\n|bench|author|petitioner|"
               r"respondent|judgment|order|coram|\Z)", " ", t, flags=re.I | re.S)
    # 5. reporter citation strings, e.g. 'AIR 2002 SC 3040', '(2002) 2 SCC 667'
    t = re.sub(rf"\(?\b(?:19|20)\d{{2}}\)?\s*\(?\d*\)?\s*(?:{REPORTERS})\.?\s*\d+",
               " ", t, flags=re.I)
    t = re.sub(rf"\b(?:{REPORTERS})\.?\s*(?:19|20)?\d{{2,4}}\s*[A-Za-z()]*\s*\d+",
               " ", t, flags=re.I)
    # 5b. residual party abbreviations and Latin / standalone reporter tags
    t = re.sub(r"\b(?:ors|anr|ano)\b\.?", " ", t, flags=re.I)         # '& Ors.', 'Anr.'
    t = re.sub(r"\(\s*(?:supra|infra)\s*\)", " ", t, flags=re.I)       # '(supra)'
    t = re.sub(r"\b(?:supra|infra|ibid)\b", " ", t, flags=re.I)
    t = re.sub(r"\b(?:SCC|SCR|AIR|ILR|MLJ|KLT|SCALE|JT|ALD|SCJ)\b", " ", t)  # caps-only reporter tags
    # 6. bench / party / metadata labels (only the all-caps LABEL: form)
    t = re.sub(r"\b(PETITIONER|RESPONDENT|APPELLANT|DATE OF JUDGMENT|BENCH|AUTHOR|"
               r"CORAM|HEADNOTE|CITATOR INFO)\s*:", " ", t)
    # 7. page numbers / running headers
    t = re.sub(r"\bpage\s+\d+\s+of\s+\d+\b", " ", t, flags=re.I)
    t = re.sub(r"\n\s*\d{1,4}\s*\n", "\n", t)
    # 8. whitespace
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{2,}", "\n", t)
    return t.strip()


def read_pdf(path: str) -> str:
    doc = fitz.open(path)
    return preprocess_legal("\n".join(pg.get_text() for pg in doc))


def extract():
    nlp = spacy.load(MODEL, disable=["ner"])
    nlp.max_length = 3_000_000
    v1.spacy_stop = nlp.Defaults.stop_words

    ids = [r["id"] for r in csv.DictReader(open(SAMPLE))
           if os.path.exists(f"{PDFDIR}/{r['id']}.pdf")]
    total = collections.Counter()
    docf = collections.Counter()
    example = {}
    print(f"Extracting noun chunks from {len(ids)} docs (legal preprocessing) ...")
    for n, did in enumerate(ids, 1):
        doc = nlp(read_pdf(f"{PDFDIR}/{did}.pdf"))
        seen = set()
        for ch in doc.noun_chunks:
            term = v1.normalise_chunk(ch)
            if not term:
                continue
            total[term] += 1
            seen.add(term)
            if term not in example:
                example[term] = re.sub(r"\s+", " ", ch.sent.text.strip())[:240]
        for t in seen:
            docf[t] += 1
        if n % 20 == 0:
            print(f"  {n}/{len(ids)} docs | unique terms: {len(total)}", flush=True)
    return total, docf, example


def frequency_adjustment(total: collections.Counter):
    """Step 5 downscale (matches codebook-reference/apply_downscale_algorithm.py).

    Each phrase is assigned to the highest-frequency unigram it contains
    (head unigram). For each unigram we discount it only by a CAPPED cumulative
    sum of its covering phrases (sorted desc), where the cap is
        max(unigram_count/2, unigram_count - max_phrase_count + 1)
    and the boundary phrase that first crosses the cap is included. This keeps
    common unigrams from being zeroed out (they retain ~half their frequency),
    unlike a full uncapped subtraction. Multi-word phrases keep their frequency.
    """
    unigram_freq = {t: f for t, f in total.items() if len(t.split()) == 1}
    assigned = collections.defaultdict(list)     # head unigram -> [(phrase, freq)]
    for term, f in total.items():
        words = term.split()
        if len(words) < 2:
            continue
        cands = [w for w in words if w in unigram_freq]
        if not cands:
            continue
        head = max(cands, key=lambda w: unigram_freq[w])
        assigned[head].append((term, f))

    adj = dict(total)
    discount = collections.Counter()
    for u, ufreq in unigram_freq.items():
        phrases = sorted(assigned.get(u, []), key=lambda x: x[1], reverse=True)
        if not phrases:
            continue
        max_phrase = phrases[0][1]
        threshold = max(ufreq / 2, ufreq - max_phrase + 1)
        cum = 0
        for _, f in phrases:
            cum += f                              # include this phrase
            if cum > threshold:                   # boundary phrase included, then stop
                break
        discount[u] = cum
        adj[u] = max(ufreq - cum, 0)
    return adj, discount


def main() -> int:
    total, docf, example = extract()
    print(f"Total unique candidate terms: {len(total)}")

    adj_freq, discount = frequency_adjustment(total)
    print(f"Unigrams discounted: {len(discount)} | "
          f"largest discounts: {discount.most_common(6)}")

    ranked = sorted(total.keys(), key=lambda t: (adj_freq[t], docf[t], total[t]),
                    reverse=True)[:TOP_N]

    rows = []
    for rank, term in enumerate(ranked, 1):
        nonamb, informative, typ, topic, role = v2.classify(term)
        rows.append({
            "rank": rank, "term": term, "doc_freq": docf[term],
            "raw_freq": total[term], "adj_freq": adj_freq[term],
            "nonambiguous": nonamb, "informative": informative,
            "type": typ, "topic": topic, "role": role,
            "notes": "", "example": example.get(term, ""),
        })
    cols = ["rank", "term", "doc_freq", "raw_freq", "adj_freq", "nonambiguous",
            "informative", "type", "topic", "role", "notes", "example"]

    with open("codebook-varshini.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(rows)
    write_xlsx(rows, cols)

    n_inf = sum(1 for r in rows if r["informative"] == "informative")
    print(f"Wrote codebook-varshini.{{csv,xlsx}} | informative: {n_inf} / "
          f"non-informative: {len(rows)-n_inf}")
    return 0


def write_xlsx(rows, cols):
    clean = lambda v: ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v
    wb = Workbook(); ws = wb.active; ws.title = "np_codebook"
    hf = PatternFill("solid", fgColor="1F4E78")
    ws.append([c.upper() for c in cols])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
    for r in rows:
        ws.append([clean(r[c]) for c in cols])
    n = len(rows)

    def dv(col, formula):
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        L = get_column_letter(cols.index(col) + 1)
        d.add(f"{L}2:{L}{n+1}")
    dv("nonambiguous", '"1,0"')
    dv("informative", '"informative,non-informative"')
    dv("type", '"' + ",".join(t for t, _ in v2.TYPES) + '"')
    dv("topic", '"' + ",".join(t for t, _ in v2.TOPICS) + '"')
    dv("role", '"' + ",".join(t for t, _ in v2.ROLES) + '"')

    widths = {"rank": 6, "term": 32, "doc_freq": 9, "raw_freq": 9, "adj_freq": 9,
              "nonambiguous": 13, "informative": 16, "type": 18, "topic": 22,
              "role": 12, "notes": 22, "example": 60}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "C2"

    # SCHEME + controlled-vocab sheets (reuse v2 vocabularies)
    sh = wb.create_sheet("SCHEME")
    for a, b in [
        ("LEGAL NOUN-PHRASE CODEBOOK — CODING PROTOCOL (v3, freq-adjusted)", ""),
        ("", ""),
        ("Ranking", "Terms are ranked by ADJ_FREQ (Step-5 adjusted frequency), not raw frequency."),
        ("adj_freq", "Unigram frequency after discounting occurrences covered by more informative phrases."),
        ("raw_freq", "Original corpus token frequency (before adjustment)."),
        ("", ""),
        ("nonambiguous (1/0)", "Stage 1: 1 = clear single legal meaning; 0 = vague/ambiguous."),
        ("informative", "Stage 2 (label): 'informative' keep / 'non-informative' drop."),
        ("type/topic/role", "Legal metadata — pick from TYPES / TOPICS / ROLES sheets."),
        ("notes / example", "Free text; example sentence for disambiguation."),
    ]:
        sh.append([a, b])
    sh["A1"].font = Font(bold=True, size=13)
    sh.column_dimensions["A"].width = 34; sh.column_dimensions["B"].width = 95

    for name, vocab in [("TYPES", v2.TYPES), ("TOPICS", v2.TOPICS), ("ROLES", v2.ROLES)]:
        vs = wb.create_sheet(name)
        vs.append([name[:-1], "DEFINITION"])
        for c in vs[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
        for v, d in vocab:
            vs.append([v, d])
        vs.column_dimensions["A"].width = 26; vs.column_dimensions["B"].width = 80

    order = ["np_codebook", "SCHEME", "TYPES", "TOPICS", "ROLES"]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    wb.save("codebook-varshini.xlsx")


if __name__ == "__main__":
    sys.exit(main())
