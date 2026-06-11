#!/usr/bin/env python3
"""
build_codebook_v2.py
====================
Upgrade codebook-varshini to the senior's multi-dimensional coding scheme,
adapted to Indian land/property-dispute law.

Reads the already-extracted top-3000 noun chunks (codebook-varshini.csv) and
enriches each term with heuristic SUGGESTIONS for:

  nonambiguous : is the term's meaning clear/unambiguous?     (1 / 0)
  keep         : final inclusion (informative)                (1 / blank)
  type         : semantic type   (controlled vocab, TYPES sheet)
  topic        : legal sub-topic (controlled vocab, TOPICS sheet)
  role         : molecule role for Step 7 (controlled vocab, ROLES sheet)

Writes codebook-varshini.xlsx with sheets:
  np_codebook  - the 3000 terms + dropdown-validated coding columns
  SCHEME       - how to code (protocol + column definitions)
  TYPES, TOPICS, ROLES - controlled vocabularies w/ definitions
and codebook-varshini.csv (same data, flat).

All coding columns are PRE-FILLED with a best-guess suggestion; the human
reviewer corrects them. This mirrors the senior's NonAmbigious/Keep/Type/Topic
layering, with his race/politics axes replaced by legal Type/Topic/Role.
"""
from __future__ import annotations
import csv, re, sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

SRC = "codebook-varshini.csv"

# ---------------- controlled vocabularies ---------------- #
TYPES = [
    ("STATUTE-PROVISION", "Statutory or constitutional reference (section, article, Act, rule, clause, proviso)"),
    ("LEGAL-CONCEPT", "Substantive doctrine, right or concept (title, possession, easement, estoppel, limitation)"),
    ("PROCEDURE-RELIEF", "Procedural act, pleading or remedy (appeal, suit, decree, writ, injunction, relief)"),
    ("INSTRUMENT-EVIDENCE", "Instrument or item of evidence (sale deed, record, register, will, mutation, map)"),
    ("PARTY-ACTOR", "Litigant, person or role in the proceedings (appellant, tenant, landlord, heir, counsel)"),
    ("FORUM-AUTHORITY", "Court, tribunal or statutory/administrative authority (High Court, Collector, Board)"),
    ("PROPERTY-RES", "The immovable property or subject-matter in dispute (land, plot, house, holding)"),
    ("JURISDICTION-PLACE", "Place or territorial jurisdiction (village, district, State, India)"),
    ("MONETARY-VALUATION", "Money, payment or valuation (compensation, consideration, rent, market value, damages)"),
    ("CITATION-REF", "Citation / reporter apparatus, non-substantive (SCC, AIR, supra, headnote, Anr)"),
    ("TEMPORAL", "Date or time expression, non-substantive (dates, months, periods)"),
    ("ASPECT", "Generic abstract noun carrying no legal content (fact, question, view, manner)"),
]
TOPICS = [
    ("TITLE-OWNERSHIP", "Title / ownership / proprietary right"),
    ("POSSESSION", "Possession / adverse possession / dispossession"),
    ("TENANCY-LEASE", "Tenancy, lease, rent, eviction, landlord-tenant"),
    ("LAND-ACQUISITION", "Compulsory acquisition / requisition / compensation"),
    ("PARTITION", "Partition / joint family / coparcenary shares"),
    ("EASEMENT", "Easement / right of way / servitude"),
    ("MORTGAGE-CHARGE", "Mortgage / charge / lien / redemption"),
    ("SUCCESSION-INHERITANCE", "Succession, inheritance, will, heirs"),
    ("SALE-CONTRACT", "Sale, agreement, conveyance, specific performance"),
    ("REVENUE-TENURE", "Zamindari/ryotwari/revenue tenure, settlement, patta"),
    ("CONSTITUTIONAL", "Constitutional/fundamental-rights dimension, writs"),
    ("PROCEDURE-REMEDY", "Procedural posture / remedy / limitation / jurisdiction"),
    ("CRIMINAL", "Criminal-law overlap (offence, accused, FIR)"),
    ("OTHER", "Does not fit the above"),
]
ROLES = [
    ("ISSUE", "The legal question/claim in dispute"),
    ("RULE", "Statute, provision, doctrine or principle applied"),
    ("EVIDENCE", "Facts, documents, records relied upon (incl. the property/res)"),
    ("ACTOR", "Parties, persons, institutions"),
    ("PROCEDURAL", "Procedural posture / metadata (appeal, decree, limitation, dates)"),
]

# ---------------- heuristic cues (suggestions only) ---------------- #
def W(*words):  # whole-word regex
    return re.compile(r"\b(" + "|".join(words) + r")\b", re.I)

ROMAN = re.compile(r"^[ivxlcdm]+$", re.I)            # iii, vii, ...
CITNUM = re.compile(r"^(no|nos|sl|vol|edn|para|pp?|cl|s|j|jj|art|sec|rw)$", re.I)

# Ordered: non-substantive noise first, generic ASPECT last; first match wins.
TYPE_CUES = [
    ("CITATION-REF", W("scc", "air", "scr", "ilr", "mlj", "klt", "all", "manu", "supra", "infra",
                       "ibid", "hereinafter", "hereinbefore", "headnote", "head note",
                       "equivalent citation", "indian kanoon", "anr", "ors", "ano", "co",
                       "ltd", "pvt", "etc", "page", "pages", "paragraph", "paragraphs",
                       "footnote", "para")),
    ("TEMPORAL", W("january", "february", "march", "april", "may", "june", "july", "august",
                   "september", "october", "november", "december", "month", "months", "year",
                   "years", "day", "days", "date", "dates", "week", "weeks", "period", "time",
                   "decade", "annum", "fortnight", "hour", "moment", "today", "yesterday",
                   "expiry", "commencement", "duration", "lapse", "interval", "deadline")),
    ("STATUTE-PROVISION", W("section", "sub-?sections?", "article", "act", "acts", "rule", "rules",
                            "clause", "sub-?clause", "schedule", "provision", "provisions",
                            "proviso", "amendment", "code", "ordinance", "statute", "regulation",
                            "regulations", "notification", "enactment", "constitution", "bye-?law",
                            "chapter", "entry", "sub-?rule", "explanation", "section\\d+", "sections")),
    ("PROCEDURE-RELIEF", W("appeal", "appeals", "special leave", "slp", "petition", "petitions",
                           "writ", "suit", "suits", "plaint", "decree", "decrees", "order",
                           "orders", "judgment", "judgement", "revision", "review", "execution",
                           "application", "proceeding", "proceedings", "hearing", "trial", "stay",
                           "injunction", "remand", "reference", "notice", "objection", "objections",
                           "plea", "pleadings?", "relief", "reliefs", "declaration", "direction",
                           "directions", "mandamus", "certiorari", "prohibition", "prayer", "motion",
                           "interlocutory", "cross-?objection", "rejoinder", "written statement",
                           "counterclaim", "summons", "warrant", "dismissal", "restoration",
                           "condonation", "appeal memo", "memo", "remedy", "remedies",
                           "interim", "ex parte", "decretal", "appellate", "procedure", "procedures",
                           "recovery", "allegation", "allegations", "demand", "service", "process",
                           "complaint", "charge-?sheet", "investigation", "inquiry", "enquiry",
                           "verdict", "acquittal", "discharge", "settlement of", "adjudication")),
    ("INSTRUMENT-EVIDENCE", W("deed", "deeds", "sale deed", "gift deed", "mortgage deed", "lease deed",
                              "document", "documents", "record", "records", "register", "registry",
                              "registration", "entry", "map", "sketch", "khasra", "jamabandi",
                              "khatauni", "patta", "receipt", "will", "wills", "testament",
                              "agreement", "agreements", "contract", "memorandum", "certificate",
                              "affidavit", "deposition", "report", "evidence", "witness", "witnesses",
                              "exhibit", "exhibits", "plan", "mutation", "endorsement", "voucher",
                              "instrument", "instruments", "letter", "letters", "correspondence")),
    ("FORUM-AUTHORITY", W("court", "courts", "supreme court", "high court", "district court",
                          "trial court", "civil court", "revenue court", "tribunal", "tribunals",
                          "bench", "board", "authority", "authorities", "commission", "government",
                          "union", "state", "corporation", "municipality", "municipal", "panchayat",
                          "collector", "magistrate", "tahsildar", "tehsildar", "registrar",
                          "sub-?registrar", "commissioner", "officer", "officers", "custodian",
                          "assessing officer", "competent authority", "appellate authority",
                          "legislature", "parliament", "judge", "judges", "judiciary", "office",
                          "department", "ministry", "secretary", "agency", "administration")),
    ("PARTY-ACTOR", W("appellant", "appellants", "respondent", "respondents", "plaintiff",
                      "plaintiffs", "defendant", "defendants", "petitioner", "petitioners",
                      "applicant", "claimant", "tenant", "tenants", "landlord", "owner", "co-?owner",
                      "proprietor", "heir", "heirs", "legal heir", "vendor", "vendee", "purchaser",
                      "buyer", "seller", "mortgagor", "mortgagee", "lessee", "lessor", "sub-?lessee",
                      "licensee", "party", "parties", "counsel", "advocate", "pleader", "occupant",
                      "trustee", "beneficiary", "guardian", "minor", "widow", "person", "persons",
                      "member", "members", "assignee", "transferee", "transferor", "decree-?holder",
                      "judgment-?debtor", "surety", "donor", "donee", "testator", "legatee",
                      "executor", "administrator", "coparcener", "karta", "ryot", "zamindar",
                      "intervener", "company", "firm", "society", "association", "trust", "appointee",
                      "wife", "husband", "son", "daughter", "father", "mother", "brother", "sister",
                      "family", "widow", "individual", "people", "public", "employee", "employer",
                      "contractor", "appointee", "nominee", "representative", "agent", "principal")),
    ("PROPERTY-RES", W("land", "lands", "property", "properties", "plot", "plots", "building",
                       "buildings", "house", "houses", "estate", "estates", "premises", "premise",
                       "tenement", "field", "fields", "well", "structure", "structures", "wall",
                       "room", "shop", "shops", "site", "holding", "holdings", "acre", "acres",
                       "bigha", "kanal", "marla", "khata", "godown", "factory", "garden", "orchard",
                       "tank", "pond", "passage", "compound", "superstructure", "construction",
                       "immovable property", "agricultural land", "boundary", "khata number")),
    ("JURISDICTION-PLACE", W("village", "villages", "mouza", "mauza", "district", "districts",
                             "tehsil", "taluka", "taluk", "city", "town", "road", "street", "india",
                             "pargana", "division", "locality", "region", "province", "ward",
                             "new delhi", "delhi", "bombay", "mumbai", "calcutta", "kolkata",
                             "madras", "chennai", "country", "countries", "nation", "territory",
                             "union territory")),
    ("MONETARY-VALUATION", W("money", "sum", "amount", "amounts", "payment", "payments",
                             "compensation", "consideration", "price", "value", "valuation", "rate",
                             "rent", "royalty", "premium", "deposit", "refund", "damages",
                             "mesne profits?", "market value", "fee", "fees", "stamp duty", "tax",
                             "taxes", "cess", "revenue", "arrears", "instalments?", "installments?",
                             "salary", "wages", "profit", "profits", "income", "damage", "debt",
                             "loan", "expense", "expenditure", "share capital")),
    ("LEGAL-CONCEPT", W("possession", "adverse possession", "title", "ownership", "right", "rights",
                        "interest", "interests", "easement", "tenancy", "lease", "mortgage",
                        "partition", "estoppel", "limitation", "jurisdiction", "acquisition",
                        "encroachment", "trespass", "succession", "inheritance", "specific performance",
                        "liability", "obligation", "covenant", "tenure", "lien", "charge", "dispute",
                        "disputes", "claim", "claims", "cause of action", "law", "laws", "principle",
                        "principles", "doctrine", "power", "powers", "validity", "vires",
                        "locus standi", "res judicata", "lis pendens", "bona fide", "mala fide",
                        "negligence", "fraud", "misrepresentation", "breach", "default", "waiver",
                        "redemption", "foreclosure", "reversion", "remainder", "equity", "good faith",
                        "natural justice", "fundamental right", "public interest", "burden of proof",
                        "onus", "presumption", "duty", "duties", "interpretation", "construction",
                        "intention", "settlement", "rights of", "right of way", "servitude", "issue",
                        "issues", "question of law", "consent", "ratification", "discretion",
                        "transfer", "sale", "purchase", "exchange", "gift", "conveyance", "assignment",
                        "adoption", "custom", "usage", "collusion", "immunity", "recovery", "merger",
                        "severance", "abandonment", "relinquishment", "acquiescence", "laches",
                        "novation", "subrogation", "indemnity", "guarantee", "partnership",
                        "dedication", "vesting", "devolution", "alienation", "validity", "legality",
                        "justice", "protection", "share", "shares", "severalty", "privilege",
                        "exemption", "concession", "grant", "franchise", "prescription")),
    ("ASPECT", W("fact", "facts", "question", "questions", "view", "views", "opinion", "circumstance",
                 "circumstances", "manner", "reason", "reasons", "matter", "matters", "case", "cases",
                 "ground", "grounds", "contention", "contentions", "submission", "submissions",
                 "aspect", "purpose", "effect", "nature", "extent", "stage", "event", "conclusion",
                 "finding", "findings", "sense", "respect", "behalf", "regard", "favour", "favor",
                 "way", "ways", "position", "hand", "word", "words", "basis", "account", "point",
                 "points", "course", "condition", "conditions", "support", "absence", "doubt",
                 "force", "light", "mind", "context", "statement", "use", "meaning", "scope",
                 "situation", "relation", "connection", "end", "virtue", "process", "observation",
                 "observations", "instance", "thing", "things", "error", "merit", "merits", "step",
                 "steps", "answer", "difference", "need", "cause", "attention", "kind", "expression",
                 "object", "form", "result", "results", "term", "terms", "place", "places", "area",
                 "areas", "portion", "part", "parts", "number", "numbers", "cost", "costs", "benefit",
                 "argument", "arguments", "declaration of", "subject", "scale", "headnote",
                 "decision", "action", "actions", "accordance", "exercise", "material", "mean",
                 "means", "determination", "reliance", "requirement", "requirements", "pursuance",
                 "substance", "operation", "control", "consequence", "consequences", "difficulty",
                 "scheme", "detail", "details", "addition", "category", "categories", "loss",
                 "attempt", "liberty", "information", "existence", "failure", "example", "examples",
                 "occasion", "controversy", "definition", "specification", "origin", "shift",
                 "clarification", "commitment", "accord", "minute", "minutes", "opposition",
                 "movement", "saving", "savings", "parity", "improvement", "guise", "total",
                 "business", "job", "death", "copy", "copies",
                 "factor", "factors", "feature", "element", "elements", "respect of",
                 "behalf of", "addition to", "difference", "comparison", "approach", "trend",
                 "tendency", "function", "role", "capacity", "status", "criterion",
                 "criteria", "standard", "level", "degree", "proportion", "percentage", "majority",
                 "minority", "balance", "outcome", "implication", "significance", "author",
                 "conduct", "language", "discussion", "life", "test", "bar", "examination",
                 "analysis", "summary", "description", "explanation of", "reference to", "type",
                 "manner of", "method", "system", "policy", "practice", "history", "background")),
]

# non-substantive types -> dropped by default in the suggestion
DROP_TYPES = {"CITATION-REF", "TEMPORAL", "ASPECT"}

TYPE_TO_ROLE = {
    "STATUTE-PROVISION": "RULE", "LEGAL-CONCEPT": "ISSUE", "PROCEDURE-RELIEF": "PROCEDURAL",
    "INSTRUMENT-EVIDENCE": "EVIDENCE", "PARTY-ACTOR": "ACTOR", "FORUM-AUTHORITY": "ACTOR",
    "PROPERTY-RES": "EVIDENCE", "JURISDICTION-PLACE": "EVIDENCE", "MONETARY-VALUATION": "EVIDENCE",
    "CITATION-REF": "", "TEMPORAL": "", "ASPECT": "",
}

TOPIC_CUES = [
    ("TENANCY-LEASE", W("tenant", "tenants", "tenancy", "lease", "lessee", "lessor", "rent",
                        "landlord", "eviction", "ejectment")),
    ("LAND-ACQUISITION", W("acquisition", "acquire", "acquired", "requisition", "land acquisition",
                           "compensation", "award", "notification")),
    ("PARTITION", W("partition", "coparcen\\w*", "joint family", "share", "shares", "karta",
                    "ancestral")),
    ("EASEMENT", W("easement", "right of way", "servitude", "passage")),
    ("MORTGAGE-CHARGE", W("mortgage", "mortgagor", "mortgagee", "redemption", "pledge", "lien",
                          "foreclosure", "charge")),
    ("SUCCESSION-INHERITANCE", W("succession", "inheritance", "heir", "heirs", "will", "wills",
                                 "legatee", "probate", "widow", "testator", "intestate")),
    ("SALE-CONTRACT", W("sale", "agreement", "specific performance", "conveyance", "vendor",
                        "vendee", "purchaser", "contract", "earnest")),
    ("TITLE-OWNERSHIP", W("title", "ownership", "owner", "proprietor", "proprietary")),
    ("POSSESSION", W("possession", "adverse possession", "occupant", "dispossession", "occupation")),
    ("REVENUE-TENURE", W("zamindar\\w*", "ryotwari", "revenue", "tenure", "patta", "khasra",
                         "jamabandi", "settlement", "ryot", "malguzar\\w*", "inam")),
    ("CONSTITUTIONAL", W("article", "fundamental right", "constitution\\w*", "writ", "ultra vires",
                         "mandamus", "certiorari")),
    ("CRIMINAL", W("offence", "accused", "criminal", "fir", "conviction", "bail", "police",
                   "prosecution", "sentence")),
    ("PROCEDURE-REMEDY", W("appeal", "suit", "decree", "injunction", "revision", "limitation",
                           "jurisdiction", "execution", "stay", "remand", "relief", "declaration")),
]


REPORTER = re.compile(r"\b(lr|l\.r|l\.\s*j|lj|cri|crilr|mad|pun|slt|srj|allcric|guj|ker|raj|"
                      r"bom|cal|del|pat|mp|ap|sc|sn|wln|fln)\b", re.I)
INITIALS = re.compile(r"\b[a-z]\.\s*[a-z]\.", re.I)
LATIN = re.compile(r"\b(inter alia|prima facie|pari materia|ratio decidendi|obiter|"
                   r"mutatis mutandis|sub silentio|ex parte|suo motu|ipso facto)\b", re.I)
FRAGMENT = {"ing", "sub", "non", "pre", "anti", "re", "de", "co", "tion", "ment", "ness"}


def classify(term: str):
    low = term.lower().strip()
    NI = "non-informative"
    # --- non-substantive noise (caught first) ---
    if ROMAN.match(low) or CITNUM.match(low) or len(low) <= 1 or low in FRAGMENT:
        return 0, NI, "CITATION-REF", "", ""
    if "- " in term:                                  # OCR hyphen-split, e.g. 'ques- tion'
        return 0, NI, "ASPECT", "", ""
    if LATIN.search(low):                             # Latin maxims used as tags
        return 0, NI, "CITATION-REF", "", ""
    if re.search(r"[()]", low) or ("&" in low and len(low) <= 14) or REPORTER.search(low):
        return 0, NI, "CITATION-REF", "", ""
    if INITIALS.search(low):                          # 's.k. das' -> person; 'l. r.' -> citation
        if re.search(r"[a-z]{4,}", low):
            return 1, "informative", "PARTY-ACTOR", "", "ACTOR"
        return 0, NI, "CITATION-REF", "", ""
    t = ""
    for name, rx in TYPE_CUES:
        if rx.search(term):
            t = name; break
    topic = ""
    for name, rx in TOPIC_CUES:
        if rx.search(term):
            topic = name; break
    role = TYPE_TO_ROLE.get(t, "")
    # non-substantive topics shouldn't tag pure-noise rows
    if t in {"CITATION-REF", "TEMPORAL"}:
        topic = ""
    nonamb = 0 if (t == "" or t in DROP_TYPES) else 1
    informative = "informative" if nonamb == 1 else "non-informative"
    return nonamb, informative, t, topic, role


def main() -> int:
    rows = list(csv.DictReader(open(SRC, encoding="utf-8")))
    out = []
    for r in rows:
        nonamb, informative, typ, topic, role = classify(r["term"])
        out.append({
            "rank": int(r["rank"]), "term": r["term"],
            "doc_freq": int(r["doc_freq"]), "total_freq": int(r["total_freq"]),
            "nonambiguous": nonamb, "informative": informative, "type": typ, "topic": topic,
            "role": role, "notes": "", "example": r.get("example", ""),
        })
    cols = ["rank", "term", "doc_freq", "total_freq", "nonambiguous", "informative",
            "type", "topic", "role", "notes", "example"]

    with open("codebook-varshini.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(out)

    write_xlsx(out, cols)

    # stats
    import collections
    for c in ["type", "topic", "role"]:
        cc = collections.Counter(r[c] or "(blank)" for r in out)
        print(f"{c}: {dict(cc.most_common())}")
    nk = sum(1 for r in out if r["informative"] == "informative")
    na = sum(1 for r in out if r["nonambiguous"] == 1)
    print(f"suggested nonambiguous=1: {na} | informative: {nk} / non-informative: {len(out)-nk}")
    print("Wrote codebook-varshini.xlsx and .csv")
    return 0


def write_xlsx(rows, cols):
    clean = lambda v: ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v
    wb = Workbook()

    # ---- main sheet ----
    ws = wb.active; ws.title = "np_codebook"
    ws.append([c.upper() for c in cols])
    hf = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
    for r in rows:
        ws.append([clean(r[c]) for c in cols])
    n = len(rows)

    def dv(col, formula):
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        L = get_column_letter(cols.index(col) + 1)
        d.add(f"{L}2:{L}{n+1}")

    dv("nonambiguous", '"1,0"')
    dv("informative", '"informative,non-informative"')
    dv("type", '"' + ",".join(t for t, _ in TYPES) + '"')
    dv("topic", '"' + ",".join(t for t, _ in TOPICS) + '"')
    dv("role", '"' + ",".join(t for t, _ in ROLES) + '"')

    widths = {"rank": 6, "term": 32, "doc_freq": 9, "total_freq": 10, "nonambiguous": 13,
              "informative": 16, "type": 18, "topic": 22, "role": 12, "notes": 22, "example": 64}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "C2"

    # ---- SCHEME sheet ----
    sh = wb.create_sheet("SCHEME")
    proto = [
        ("LEGAL NOUN-PHRASE CODEBOOK — CODING PROTOCOL", ""),
        ("", ""),
        ("Goal", "For each of the top 3000 noun-chunk terms, record the coding columns below."),
        ("", "Coding columns are PRE-FILLED with a suggestion; review and correct each."),
        ("", ""),
        ("COLUMN", "DEFINITION"),
        ("rank/term/doc_freq/total_freq", "Auto: term and its frequency across the 100 documents."),
        ("nonambiguous (1/0)", "Stage 1: 1 = the term has a clear, single legal meaning. 0 = vague/generic/ambiguous."),
        ("informative", "Stage 2 (the label): 'informative' = keep as a meaningful legal term; 'non-informative' = drop as boilerplate/noise."),
        ("type", "Semantic type of the term — pick from the TYPES sheet."),
        ("topic", "Legal sub-topic the term belongs to — pick from the TOPICS sheet (blank if none)."),
        ("role", "Molecule role for Step 7 — pick from the ROLES sheet (issue/rule/evidence/actor/procedural)."),
        ("notes", "Free text: merges, doubts, alternate readings."),
        ("example", "A real sentence the term appeared in — use for disambiguation."),
        ("", ""),
        ("TWO-STAGE DECISION", "First decide nonambiguous; only then decide keep. A term can be unambiguous but still dropped if not useful."),
        ("DROPDOWNS", "type / topic / role / nonambiguous / keep are dropdown-validated to the controlled vocabularies."),
    ]
    for a, b in proto:
        sh.append([a, b])
    sh["A1"].font = Font(bold=True, size=13)
    sh["A6"].font = sh["B6"].font = Font(bold=True)
    sh.column_dimensions["A"].width = 34; sh.column_dimensions["B"].width = 95

    # ---- vocab sheets ----
    for name, vocab in [("TYPES", TYPES), ("TOPICS", TOPICS), ("ROLES", ROLES)]:
        vs = wb.create_sheet(name)
        vs.append([name[:-1] if name.endswith("S") else name, "DEFINITION"])
        for c in vs[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
        for v, d in vocab:
            vs.append([v, d])
        vs.column_dimensions["A"].width = 26; vs.column_dimensions["B"].width = 80

    # order sheets: main, SCHEME, TYPES, TOPICS, ROLES
    wb._sheets.sort(key=lambda s: ["np_codebook", "SCHEME", "TYPES", "TOPICS", "ROLES"].index(s.title))
    wb.save("codebook-varshini.xlsx")


if __name__ == "__main__":
    sys.exit(main())
