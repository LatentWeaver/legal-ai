#!/usr/bin/env python3
"""
build_np_codebook.py
====================
Build the NOUN-CHUNK ("np") legal codebook for the 100 sampled documents.

Pipeline:
  1. Read sample_100_docs.csv -> 100 doc ids.
  2. Extract + clean text from documents_pdf/<id>.pdf  (PyMuPDF).
  3. spaCy noun_chunks over each doc; normalise each chunk
     (drop leading determiners/stopwords, lemmatise, lowercase).
  4. Tally document-frequency and total-frequency; keep one example.
  5. Rank by doc-freq (tie: total-freq); take top 3000.
  6. Write:
       codebook-varshini.csv
       codebook-varshini.xlsx   (with a label dropdown: informative/non-informative)

A `suggested_label` helper column is pre-filled with a heuristic guess so the
manual reviewer can sort/accept quickly; the `label` column is left blank for
the human's final, manual decision.
"""
from __future__ import annotations
import csv, os, re, sys, collections
import fitz  # PyMuPDF
import spacy

SAMPLE = "sample_100_docs.csv"
PDFDIR = "documents_pdf"
TOP_N = 3000
MODEL = "en_core_web_md"

# ----- boilerplate / non-informative heuristic cues (for suggested_label only) -----
BOILERPLATE = {
    "court", "learned counsel", "counsel", "appeal", "respondent", "appellant",
    "petitioner", "judgment", "order", "case", "matter", "fact", "question",
    "view", "manner", "behalf", "regard", "respect", "ground", "contention",
    "submission", "para", "paragraph", "page", "date", "day", "year", "time",
    "present case", "instant case", "high court", "supreme court", "trial court",
    "learned judge", "hon ble court", "honourable court", "writ petition",
    "special leave petition", "civil appeal", "criminal appeal",
}
LEGAL_CUE = re.compile(
    r"\b(section|article|act|rule|order|clause|schedule|amendment|provision|"
    r"right|title|possession|easement|tenancy|lease|mortgage|partition|"
    r"decree|injunction|estoppel|limitation|jurisdiction|acquisition|"
    r"encroachment|trespass|conveyance|sale deed|gift deed|will|succession|"
    r"zamindari|ryotwari|revenue|registration|specific performance)\b",
    re.I,
)


def clean_text(raw: str) -> str:
    # Collapse PDF letter-spacing artifacts: runs of >=4 single letters w/ spaces.
    raw = re.sub(r"(?:\b[A-Za-z]\b ){3,}\b[A-Za-z]\b",
                 lambda m: m.group(0).replace(" ", ""), raw)
    raw = raw.replace("\xad", "")
    raw = re.sub(r"[ \t]+", " ", raw)
    raw = re.sub(r"\n{2,}", "\n", raw)
    return raw.strip()


def pdf_text(path: str) -> str:
    doc = fitz.open(path)
    return clean_text("\n".join(pg.get_text() for pg in doc))


def normalise_chunk(chunk) -> str | None:
    toks = list(chunk)
    # drop leading determiners / pronouns / possessives / stopwords
    while toks and (toks[0].pos_ in {"DET", "PRON"} or toks[0].is_stop
                    or toks[0].is_punct or toks[0].is_space):
        toks = toks[1:]
    # drop trailing punctuation/space
    while toks and (toks[-1].is_punct or toks[-1].is_space):
        toks = toks[:-1]
    if not toks or len(toks) > 5:
        return None
    parts = []
    for t in toks:
        if t.is_punct or t.is_space or t.like_num:
            continue
        lemma = t.lemma_.lower().strip()
        if lemma:
            parts.append(lemma)
    if not parts:
        return None
    term = " ".join(parts)
    term = re.sub(r"\s+", " ", term).strip(" -")
    # filters: drop single chars, pure non-alpha, all-stopword remnants
    if len(term) < 3 or not re.search(r"[a-z]", term):
        return None
    if all(w in spacy_stop for w in term.split()):
        return None
    return term


def suggest_label(term: str) -> str:
    if term in BOILERPLATE:
        return "non-informative"
    if LEGAL_CUE.search(term):
        return "informative"
    # multi-word domain phrases lean informative; bare generic single words lean non
    return "informative" if len(term.split()) >= 2 else "non-informative"


def main() -> int:
    if not os.path.exists(SAMPLE):
        print("missing", SAMPLE); return 2
    ids = [r["id"] for r in csv.DictReader(open(SAMPLE))]
    missing = [i for i in ids if not os.path.exists(f"{PDFDIR}/{i}.pdf")]
    if missing:
        print(f"WARNING: {len(missing)} PDFs missing, proceeding with {len(ids)-len(missing)}")
    ids = [i for i in ids if os.path.exists(f"{PDFDIR}/{i}.pdf")]

    print(f"Loading {MODEL} ...")
    nlp = spacy.load(MODEL, disable=["ner"])
    nlp.max_length = 3_000_000
    global spacy_stop
    spacy_stop = nlp.Defaults.stop_words

    total_freq = collections.Counter()
    doc_freq = collections.Counter()
    example = {}

    for n, did in enumerate(ids, 1):
        text = pdf_text(f"{PDFDIR}/{did}.pdf")
        doc = nlp(text)
        seen_in_doc = set()
        for chunk in doc.noun_chunks:
            term = normalise_chunk(chunk)
            if not term:
                continue
            total_freq[term] += 1
            seen_in_doc.add(term)
            if term not in example:
                sent = chunk.sent.text.strip().replace("\n", " ")
                example[term] = re.sub(r"\s+", " ", sent)[:240]
        for term in seen_in_doc:
            doc_freq[term] += 1
        if n % 10 == 0:
            print(f"  processed {n}/{len(ids)} docs | unique terms so far: {len(total_freq)}", flush=True)

    print(f"Total unique noun-chunk terms: {len(total_freq)}")
    ranked = sorted(total_freq.keys(),
                    key=lambda t: (doc_freq[t], total_freq[t]), reverse=True)[:TOP_N]

    # ---- write CSV ----
    rows = []
    for rank, term in enumerate(ranked, 1):
        rows.append({
            "rank": rank, "term": term,
            "doc_freq": doc_freq[term], "total_freq": total_freq[term],
            "suggested_label": suggest_label(term),
            "label": "", "notes": "",
            "example": example.get(term, ""),
        })
    cols = ["rank", "term", "doc_freq", "total_freq", "suggested_label", "label", "notes", "example"]
    with open("codebook-varshini.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(rows)
    print("Wrote codebook-varshini.csv")

    # ---- write XLSX with label dropdown ----
    write_xlsx(rows, cols)
    n_inf = sum(1 for r in rows if r["suggested_label"] == "informative")
    print(f"Wrote codebook-varshini.xlsx  | suggested: {n_inf} informative / {len(rows)-n_inf} non-informative")
    return 0


def write_xlsx(rows, cols):
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    def clean(v):
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v

    wb = Workbook(); ws = wb.active; ws.title = "np_codebook"
    ws.append([c.upper() for c in cols])
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = head_fill
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([clean(r[c]) for c in cols])
    dv = DataValidation(type="list", formula1='"informative,non-informative"', allow_blank=True)
    ws.add_data_validation(dv)
    label_col = cols.index("label") + 1
    from openpyxl.utils import get_column_letter
    L = get_column_letter(label_col)
    dv.add(f"{L}2:{L}{len(rows)+1}")
    widths = {"rank": 6, "term": 34, "doc_freq": 9, "total_freq": 10,
              "suggested_label": 16, "label": 16, "notes": 24, "example": 70}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "A2"
    wb.save("codebook-varshini.xlsx")


if __name__ == "__main__":
    sys.exit(main())
