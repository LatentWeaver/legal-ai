#!/usr/bin/env python3
"""
build_codebook_v4.py
====================
Workflow Step-2 compliant: MULTI-METHOD candidate extraction.

Candidate phrases are pooled from four extractors over the 100
legally-preprocessed documents:
  - spaCy   noun_chunks            (syntactic noun phrases)
  - Textacy keyterms.textrank      (graph-based keyphrases)
  - RAKE    rake_nltk              (co-occurrence keyphrases)
  - YAKE    yake                   (statistical keyphrases)
(Named entities are the separate NE codebook — Vibha's half.)

All candidates are normalised to a lemma key. A single uniform corpus
frequency is then computed for EVERY candidate by a lemma n-gram sweep
(Step 4), so methods that return ranked sets (RAKE/YAKE/Textacy) and
methods that return spans (spaCy) are put on the same footing. Then:
  Step 5  frequency adjustment of common unigrams (import from v3)
  Step 6  rank by adjusted frequency
  Step 7  top 3000
  Step 8/9 pre-annotate (informative/type/topic/role) via v2 heuristic

A `methods` column records which extractors proposed each term.
Outputs codebook-varshini.{csv,xlsx}.
"""
from __future__ import annotations
import csv, os, re, collections, sys
import ssl
try:
    ssl._create_default_https_context = ssl._create_unverified_context
except AttributeError:
    pass

import spacy
import yake
from rake_nltk import Rake
from textacy.extract import keyterms as tk

import build_np_codebook as v1     # normalise_chunk
import build_codebook_v2 as v2     # classify, TYPES/TOPICS/ROLES
import build_codebook_v3 as v3     # read_pdf (legal preprocessing), frequency_adjustment
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

SAMPLE, PDFDIR, TOP_N, MODEL, MAXLEN = "sample_100_docs.csv", "documents_pdf", 3000, "en_core_web_md", 5
PER_DOC = 25   # keyphrases kept per document from each ranked extractor


DET_EXTRA = {"the", "a", "an", "this", "that", "these", "those", "said", "such",
             "its", "their", "his", "her", "our", "your", "any", "each", "every"}


def strip_bounds(term, STOP):
    """String-based leading/trailing determiner+stopword strip (handles ALL-CAPS
    headers that spaCy mis-tags, e.g. 'THE APPELLANT' -> 'appellant')."""
    if not term:
        return None
    bound = STOP | DET_EXTRA
    toks = term.split()
    while toks and toks[0] in bound:
        toks = toks[1:]
    while toks and toks[-1] in bound:
        toks = toks[:-1]
    if not toks or len(toks) > MAXLEN:
        return None
    out = " ".join(toks)
    return out if len(out) >= 3 else None


def normalise_lemmas(toks, STOP):
    while toks and (toks[0] in STOP or not re.search(r"[a-z]", toks[0])):
        toks = toks[1:]
    while toks and (toks[-1] in STOP or not re.search(r"[a-z]", toks[-1])):
        toks = toks[:-1]
    toks = [w for w in toks if re.search(r"[a-z]", w)]
    if not toks or len(toks) > MAXLEN:
        return None
    term = " ".join(toks).strip()
    return term if len(term) >= 3 else None


def main() -> int:
    nlp = spacy.load(MODEL, disable=["ner"])      # parser needed for noun_chunks
    nlp.max_length = 3_000_000
    v1.spacy_stop = nlp.Defaults.stop_words
    STOP = nlp.Defaults.stop_words
    light = spacy.load(MODEL, disable=["parser", "ner"])   # lemma-only, for RAKE/YAKE strings
    rake = Rake(max_length=MAXLEN)
    yk = yake.KeywordExtractor(lan="en", n=MAXLEN, top=PER_DOC, dedupLim=0.9)

    ids = [r["id"] for r in csv.DictReader(open(SAMPLE))
           if os.path.exists(f"{PDFDIR}/{r['id']}.pdf")]

    CAND = set()
    methods = collections.defaultdict(set)
    example = {}
    doc_lemmas = []
    rake_raw, yake_raw = set(), set()

    print(f"Extracting candidates from {len(ids)} docs via spaCy+Textacy+RAKE+YAKE ...")
    for n, did in enumerate(ids, 1):
        text = v3.read_pdf(f"{PDFDIR}/{did}.pdf")
        doc = nlp(text)
        doc_lemmas.append([t.lemma_.lower() for t in doc if not (t.is_punct or t.is_space)])

        for ch in doc.noun_chunks:                       # 1. spaCy
            key = strip_bounds(v1.normalise_chunk(ch), STOP)
            if key:
                CAND.add(key); methods[key].add("spacy")
                example.setdefault(key, re.sub(r"\s+", " ", ch.sent.text.strip())[:240])
        try:                                             # 2. Textacy
            for term, _ in tk.textrank(doc, normalize="lemma", topn=PER_DOC):
                key = normalise_lemmas(term.split(), STOP)
                if key:
                    CAND.add(key); methods[key].add("textacy")
        except Exception:
            pass
        try:                                             # 3. RAKE
            rake.extract_keywords_from_text(text)
            rake_raw.update(rake.get_ranked_phrases()[:PER_DOC])
        except Exception:
            pass
        try:                                             # 4. YAKE
            yake_raw.update(kw for kw, _ in yk.extract_keywords(text))
        except Exception:
            pass
        if n % 20 == 0:
            print(f"  {n}/{len(ids)} docs | candidates so far: {len(CAND)}", flush=True)

    # lemmatise RAKE / YAKE surface phrases once, in bulk
    for raw_set, name in [(rake_raw, "rake"), (yake_raw, "yake")]:
        raws = list(raw_set)
        for d in light.pipe(raws, batch_size=256):
            key = normalise_lemmas([t.lemma_.lower() for t in d if not (t.is_punct or t.is_space)], STOP)
            if key:
                CAND.add(key); methods[key].add(name)
    print(f"Total unique candidates (all methods merged): {len(CAND)}")

    # ---- Step 4: uniform corpus frequency via lemma n-gram sweep ----
    total = collections.Counter()
    docf = collections.Counter()
    for lem in doc_lemmas:
        seen = set()
        L = len(lem)
        for i in range(L):
            for nlen in range(1, MAXLEN + 1):
                if i + nlen > L:
                    break
                if lem[i] in STOP or lem[i + nlen - 1] in STOP:
                    continue                              # candidates never start/end on a stopword
                gram = " ".join(lem[i:i + nlen])
                if gram in CAND:
                    total[gram] += 1
                    seen.add(gram)
        for g in seen:
            docf[g] += 1
    # candidates never actually found in text get freq 0 -> drop
    terms = [t for t in CAND if total[t] > 0]
    print(f"Candidates with corpus freq > 0: {len(terms)}")

    # ---- noun-phrase filter: this is the NP codebook, so drop verb/adjective/
    # adverb candidates that RAKE/YAKE/Textacy emit. Keep a term if its head
    # (last) token is a NOUN/PROPN, OR it is a recognised legal term. ----
    SUBSTANTIVE = {"STATUTE-PROVISION", "LEGAL-CONCEPT", "PROCEDURE-RELIEF",
                   "INSTRUMENT-EVIDENCE", "PARTY-ACTOR", "FORUM-AUTHORITY",
                   "PROPERTY-RES", "JURISDICTION-PLACE", "MONETARY-VALUATION"}
    noun_headed = set()
    tl = list(terms)
    for term, d in zip(tl, light.pipe(tl, batch_size=512)):
        toks = [t for t in d if not (t.is_punct or t.is_space)]
        if toks and toks[-1].pos_ in {"NOUN", "PROPN"}:
            noun_headed.add(term)
    before = len(terms)
    terms = [t for t in terms if t in noun_headed or v2.classify(t)[2] in SUBSTANTIVE]
    print(f"After noun-phrase filter: {len(terms)} (dropped {before - len(terms)} non-noun candidates)")

    # ---- Step 5/6: frequency adjustment + ranking ----
    adj_freq, discount = v3.frequency_adjustment(collections.Counter({t: total[t] for t in terms}))
    ranked = sorted(terms, key=lambda t: (adj_freq[t], docf[t], total[t]), reverse=True)[:TOP_N]

    rows = []
    for rank, term in enumerate(ranked, 1):
        nonamb, informative, typ, topic, role = v2.classify(term)
        rows.append({
            "rank": rank, "term": term, "doc_freq": docf[term],
            "raw_freq": total[term], "adj_freq": adj_freq[term],
            "methods": ",".join(sorted(methods[term])),
            "nonambiguous": nonamb, "informative": informative,
            "type": typ, "topic": topic, "role": role,
            "notes": "", "example": example.get(term, ""),
        })
    cols = ["rank", "term", "doc_freq", "raw_freq", "adj_freq", "methods",
            "nonambiguous", "informative", "type", "topic", "role", "notes", "example"]

    with open("codebook-varshini.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols); w.writeheader(); w.writerows(rows)
    write_xlsx(rows, cols)

    # ---- method contribution report ----
    from collections import Counter
    method_terms = Counter()
    multi = 0
    for r in rows:
        ms = r["methods"].split(",")
        for m in ms:
            method_terms[m] += 1
        if len(ms) > 1:
            multi += 1
    n_inf = sum(1 for r in rows if r["informative"] == "informative")
    print(f"\nMethod contribution within top {TOP_N}:")
    for m, c in method_terms.most_common():
        print(f"  {m:<9} {c}")
    print(f"  multi-method (>=2): {multi}")
    print(f"informative: {n_inf} / non-informative: {len(rows) - n_inf}")
    print("Wrote codebook-varshini.{csv,xlsx}")
    return 0


def write_xlsx(rows, cols):
    clean = lambda v: ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v
    wb = Workbook(); ws = wb.active; ws.title = "np_codebook"
    hf = PatternFill("solid", fgColor="1F4E78")
    ws.append([c.upper() for c in cols])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
    for r in rows:
        ws.append([clean(r[c]) for c in cols])
    n = len(rows)

    def dv(col, formula):
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        L = get_column_letter(cols.index(col) + 1)
        d.add(f"{L}2:{L}{n+1}")
    dv("nonambiguous", '"1,0"')
    dv("informative", '"informative,non-informative"')
    dv("type", '"' + ",".join(t for t, _ in v2.TYPES) + '"')
    dv("topic", '"' + ",".join(t for t, _ in v2.TOPICS) + '"')
    dv("role", '"' + ",".join(t for t, _ in v2.ROLES) + '"')

    widths = {"rank": 6, "term": 30, "doc_freq": 9, "raw_freq": 9, "adj_freq": 9,
              "methods": 20, "nonambiguous": 13, "informative": 16, "type": 18,
              "topic": 22, "role": 12, "notes": 20, "example": 55}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 14)
    ws.freeze_panes = "C2"

    sh = wb.create_sheet("SCHEME")
    for a, b in [
        ("LEGAL NP CODEBOOK — multi-method (v4)", ""),
        ("", ""),
        ("methods", "Which extractors proposed the term: spacy / textacy / rake / yake."),
        ("adj_freq", "Step-5 adjusted frequency (ranking key); raw_freq = pre-adjustment."),
        ("nonambiguous (1/0)", "Stage 1: clear single legal meaning?"),
        ("informative", "Stage 2 label: informative / non-informative."),
        ("type/topic/role", "Legal metadata — see TYPES / TOPICS / ROLES sheets."),
    ]:
        sh.append([a, b])
    sh["A1"].font = Font(bold=True, size=13)
    sh.column_dimensions["A"].width = 32; sh.column_dimensions["B"].width = 92
    for name, vocab in [("TYPES", v2.TYPES), ("TOPICS", v2.TOPICS), ("ROLES", v2.ROLES)]:
        vs = wb.create_sheet(name)
        vs.append([name[:-1], "DEFINITION"])
        for c in vs[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = hf
        for v, d in vocab:
            vs.append([v, d])
        vs.column_dimensions["A"].width = 26; vs.column_dimensions["B"].width = 80
    order = ["np_codebook", "SCHEME", "TYPES", "TOPICS", "ROLES"]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    wb.save("codebook-varshini.xlsx")


if __name__ == "__main__":
    sys.exit(main())
